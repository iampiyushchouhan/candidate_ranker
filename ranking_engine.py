"""
Neural Ranking Engine — 3-Stage Pipeline
Stage 1: Pre-computed embeddings (via precompute.py)
Stage 2: Dense semantic retrieval (cosine similarity → top-K shortlist)
Stage 3: Neural re-ranking (structured features + semantic score → final top 100)
"""

import os
import sys
import json
import time
import csv
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import (
    JOB_DESCRIPTION, RERANK_WEIGHTS, RETRIEVAL_TOP_K, FINAL_TOP_N,
)
from scoring.semantic_scorer import (
    encode_texts, encode_single, build_candidate_text,
    cosine_similarity_batch, retrieve_top_k,
    load_embeddings, save_embeddings,
)
from scoring.feature_extractor import extract_features
from scoring.disqualifiers import apply_disqualifiers
from scoring.honeypot_detector import detect_honeypot


def _compute_rerank_score(features: dict, semantic_sim: float) -> float:
    """
    Compute the re-ranking score from structured features + semantic similarity.
    This is a hand-tuned weighted combination (acts like a lightweight neural scorer).
    """
    w = RERANK_WEIGHTS

    # --- Semantic similarity (0-1, already computed) ---
    sem_score = max(0.0, min(1.0, semantic_sim))

    # --- Skills fit (0-1) ---
    req_cov = min(1.0, features.get("required_skill_coverage", 0))
    pref_cov = min(1.0, features.get("preferred_skill_coverage", 0))
    ai_ratio = features.get("ai_skill_ratio", 0)
    nlp_ir = min(1.0, features.get("nlp_ir_skill_count", 0) / 5.0)  # Up to 5 NLP/IR skills = 1.0
    prof = features.get("avg_ai_proficiency", 0)
    assess = features.get("avg_assessment_score", 0) / 100.0
    has_assess = features.get("has_assessments", 0)

    skills_score = (
        req_cov * 0.25 +
        pref_cov * 0.10 +
        ai_ratio * 0.15 +
        nlp_ir * 0.20 +
        prof * 0.15 +
        (assess * 0.15 if has_assess else prof * 0.15)
    )

    # --- Career trajectory (0-1) ---
    title_relevant = features.get("title_is_relevant", 0)
    title_ai = features.get("title_has_ai_keyword", 0)
    career_ai_ratio = features.get("career_ai_title_ratio", 0)
    desc_ml_ratio = features.get("desc_ml_ratio", 0)
    has_product = features.get("has_product_experience", 0)
    stability = 0.0
    avg_tenure = features.get("avg_tenure_months", 0)
    if avg_tenure >= 30:
        stability = 1.0
    elif avg_tenure >= 24:
        stability = 0.8
    elif avg_tenure >= 18:
        stability = 0.6
    elif avg_tenure >= 12:
        stability = 0.4
    else:
        stability = 0.2

    career_score = (
        title_relevant * 0.15 +
        title_ai * 0.15 +
        career_ai_ratio * 0.20 +
        desc_ml_ratio * 0.20 +
        has_product * 0.15 +
        stability * 0.15
    )

    # --- Behavioral signals (0-1) ---
    response_rate = features.get("recruiter_response_rate", 0)
    recently_active = features.get("recently_active", 0)
    github = min(1.0, features.get("github_activity", 0) / 80.0)
    notice_ok = features.get("notice_under_30", 0) * 0.5 + features.get("notice_under_60", 0) * 0.3
    interview_rate = features.get("interview_completion_rate", 0)
    completeness = features.get("profile_completeness", 0)
    open_to_work = features.get("open_to_work", 0)

    behavioral_score = (
        response_rate * 0.25 +
        recently_active * 0.15 +
        github * 0.15 +
        notice_ok * 0.10 +
        interview_rate * 0.10 +
        completeness * 0.10 +
        open_to_work * 0.10 +
        features.get("verification_score", 0) * 0.05
    )

    # --- Experience fit (0-1) ---
    exp_ideal = features.get("exp_in_ideal_range", 0)
    exp_ok = features.get("exp_in_acceptable_range", 0)
    exp_distance = features.get("exp_distance", 0)
    exp_penalty = max(0, 1.0 - exp_distance * 0.1)

    experience_score = exp_ideal * 0.5 + exp_ok * 0.3 + exp_penalty * 0.2

    # --- Location (added as a bonus multiplier) ---
    location_score = features.get("location_score", 0.5)

    # --- Education (0-1) ---
    edu_score = (
        features.get("edu_tier_score", 0.3) * 0.3 +
        features.get("edu_degree_score", 0.4) * 0.3 +
        features.get("edu_field_relevant", 0) * 0.25 +
        min(1.0, features.get("relevant_cert_count", 0) / 2.0) * 0.15
    )

    # --- Combine with weights ---
    raw_score = (
        sem_score * w["semantic_similarity"] +
        skills_score * w["skills_fit"] +
        career_score * w["career_trajectory"] +
        behavioral_score * w["behavioral_signals"] +
        experience_score * w["experience_fit"] +
        edu_score * w["education"]
    )

    # Apply location as a soft multiplier (0.7x - 1.0x)
    location_multiplier = 0.7 + location_score * 0.3
    raw_score *= location_multiplier

    return max(0.0, min(1.0, raw_score))


def rank_candidates(
    candidates: list,
    embeddings_dir: str = None,
    progress_callback=None,
    top_k: int = RETRIEVAL_TOP_K,
    top_n: int = FINAL_TOP_N,
) -> list:
    """
    Full 3-stage ranking pipeline.

    Args:
        candidates: List of candidate dicts
        embeddings_dir: Directory with pre-computed embeddings (optional)
        progress_callback: fn(current, total, stage, elapsed)
        top_k: Shortlist size from semantic retrieval
        top_n: Final output size

    Returns:
        List of ranked result dicts (best first)
    """
    start_time = time.time()
    total = len(candidates)

    # =====================================================================
    # STAGE 2: Dense Semantic Retrieval
    # =====================================================================
    if progress_callback:
        progress_callback(0, total, "Loading embeddings...", 0)

    # Try to load pre-computed embeddings
    candidate_embeddings = None
    jd_embedding = None
    candidate_ids_map = {}

    if embeddings_dir:
        emb_path = os.path.join(embeddings_dir, "candidate_embeddings.npy")
        jd_path = os.path.join(embeddings_dir, "jd_embedding.npy")
        ids_path = os.path.join(embeddings_dir, "candidate_ids.json")

        candidate_embeddings = load_embeddings(emb_path)
        jd_embedding = load_embeddings(jd_path)

        if os.path.exists(ids_path):
            with open(ids_path, "r") as f:
                stored_ids = json.load(f)
            candidate_ids_map = {cid: i for i, cid in enumerate(stored_ids)}

    # If no pre-computed embeddings, compute on the fly
    if candidate_embeddings is None:
        if progress_callback:
            progress_callback(0, total, "Computing embeddings on-the-fly...", time.time() - start_time)

        texts = []
        for i, c in enumerate(candidates):
            texts.append(build_candidate_text(c))
            if progress_callback and (i + 1) % 1000 == 0:
                progress_callback(i + 1, total, "Building text profiles...", time.time() - start_time)

        if progress_callback:
            progress_callback(total, total, "Encoding with sentence-transformers...", time.time() - start_time)

        candidate_embeddings = encode_texts(texts, batch_size=256, show_progress=True)

    if jd_embedding is None:
        jd_text = JOB_DESCRIPTION.get("raw_text", "")
        jd_embedding = encode_single(jd_text)

    # Retrieve top-K by semantic similarity
    if progress_callback:
        progress_callback(0, top_k, "Dense retrieval (cosine search)...", time.time() - start_time)

    # Match candidates to embedding indices
    if candidate_ids_map:
        # Pre-computed: need to map candidate IDs to embedding indices
        all_sims = cosine_similarity_batch(jd_embedding, candidate_embeddings)
        top_indices = np.argsort(all_sims)[::-1][:top_k]
        top_sim_scores = all_sims[top_indices]

        # Map back to candidates
        stored_ids_list = list(candidate_ids_map.keys())
        shortlisted = []
        for idx, sim in zip(top_indices, top_sim_scores):
            cid = stored_ids_list[idx] if idx < len(stored_ids_list) else None
            # Find candidate by ID
            for c in candidates:
                if c.get("candidate_id") == cid:
                    shortlisted.append((c, float(sim)))
                    break
    else:
        # On-the-fly: indices match directly
        all_sims = cosine_similarity_batch(jd_embedding, candidate_embeddings)
        top_indices = np.argsort(all_sims)[::-1][:top_k]
        top_sim_scores = all_sims[top_indices]
        shortlisted = [(candidates[idx], float(sim)) for idx, sim in zip(top_indices, top_sim_scores)]

    retrieval_time = time.time() - start_time
    print(f"\n  Stage 2 complete: {len(shortlisted)} candidates shortlisted in {retrieval_time:.1f}s")
    if shortlisted:
        print(f"  Semantic sim range: {shortlisted[0][1]:.4f} - {shortlisted[-1][1]:.4f}")

    # =====================================================================
    # STAGE 3: Neural Re-Ranking
    # =====================================================================
    if progress_callback:
        progress_callback(0, len(shortlisted), "Re-ranking with structured features...", time.time() - start_time)

    results = []

    for i, (candidate, sem_sim) in enumerate(shortlisted):
        try:
            # Extract structured features
            features = extract_features(candidate)

            # Compute re-rank score
            rerank_score = _compute_rerank_score(features, sem_sim)

            # Apply disqualifiers
            disq = apply_disqualifiers(candidate, features)

            # Apply honeypot penalty
            hp = detect_honeypot(candidate, features)

            # Final score = rerank × disqualifier × honeypot
            final_score = rerank_score * disq["multiplier"] * hp["penalty"]

            profile = candidate.get("profile", {})

            # Build reasoning
            reasoning = _build_reasoning(candidate, features, sem_sim, rerank_score, disq, hp)

            results.append({
                "candidate_id": candidate.get("candidate_id", ""),
                "rank": 0,  # Assigned later
                "final_score": round(final_score, 6),
                "raw_rerank_score": round(rerank_score, 6),
                "semantic_similarity": round(sem_sim, 6),
                "disqualifier_multiplier": round(disq["multiplier"], 4),
                "honeypot_penalty": round(hp["penalty"], 4),
                "reasoning": reasoning,
                # Profile info for display
                "name": profile.get("anonymized_name", ""),
                "current_title": profile.get("current_title", ""),
                "years_experience": profile.get("years_of_experience", 0),
                "location": profile.get("location", ""),
                "country": profile.get("country", ""),
                "company": profile.get("current_company", ""),
                "industry": profile.get("current_industry", ""),
                # Detail breakdowns
                "features": features,
                "disqualifier_reasons": disq["reasons"],
                "is_disqualified": disq["is_disqualified"],
                "honeypot_flags": hp["flags"],
                "is_honeypot": hp["is_honeypot"],
            })

        except Exception as e:
            print(f"  Error scoring {candidate.get('candidate_id', '?')}: {e}")

        if progress_callback and (i + 1) % 50 == 0:
            progress_callback(i + 1, len(shortlisted), "Re-ranking...", time.time() - start_time)

    # Sort by final score descending, tie-break by candidate_id
    results.sort(key=lambda x: (-x["final_score"], x["candidate_id"]))

    # Assign ranks
    for i, r in enumerate(results):
        r["rank"] = i + 1

    total_time = time.time() - start_time
    print(f"  Stage 3 complete: {len(results)} candidates scored in {total_time:.1f}s")

    honeypots_in_top100 = sum(1 for r in results[:100] if r.get("is_honeypot", False))
    disqualified_in_top100 = sum(1 for r in results[:100] if r.get("is_disqualified", False))
    print(f"  Honeypots in top 100: {honeypots_in_top100}")
    print(f"  Disqualified in top 100: {disqualified_in_top100}")

    return results


def _build_reasoning(candidate: dict, features: dict, sem_sim: float,
                     rerank_score: float, disq: dict, hp: dict) -> str:
    """Build a natural-language reasoning string for the submission CSV."""
    profile = candidate.get("profile", {})
    parts = []

    title = profile.get("current_title", "Unknown")
    years = profile.get("years_of_experience", 0)
    company = profile.get("current_company", "")
    location = profile.get("location", "")

    # Core match
    parts.append(f"{title} with {years:.0f} years experience")
    if company:
        parts.append(f"at {company}")

    # NLP/IR fit
    nlp_count = features.get("nlp_ir_skill_count", 0)
    if nlp_count >= 3:
        parts.append(f"strong NLP/IR skills ({nlp_count} relevant)")
    elif nlp_count >= 1:
        parts.append(f"some NLP/IR exposure ({nlp_count} relevant)")

    # Required skill coverage
    req_matched = features.get("required_skills_matched", 0)
    if req_matched >= 8:
        parts.append(f"high required-skill coverage ({req_matched} matched)")
    elif req_matched >= 4:
        parts.append(f"moderate required-skill coverage ({req_matched} matched)")

    # Product vs consulting
    if features.get("has_product_experience", 0) == 1.0:
        parts.append("product company experience")
    if features.get("is_consulting_only", 0) == 1.0:
        parts.append("consulting-only career")

    # Location
    if location:
        parts.append(f"based in {location}")

    # Behavioral
    rr = features.get("recruiter_response_rate", 0)
    if rr >= 0.7:
        parts.append(f"strong engagement ({rr:.0%} response rate)")
    elif rr >= 0.4:
        parts.append(f"moderate engagement ({rr:.0%} response rate)")
    elif rr < 0.2:
        parts.append(f"low recruiter response rate ({rr:.0%})")

    # Flags
    if hp.get("is_honeypot", False):
        parts.append("⚠️ possible honeypot")
    if disq.get("reasons"):
        parts.append(f"concerns: {disq['reasons'][0]}")

    notice = features.get("notice_period_days", 90)
    if notice <= 30:
        parts.append(f"available within {notice} days")
    elif notice > 90:
        parts.append(f"long notice period ({notice} days)")

    return "; ".join(parts) + "."


def get_top_candidates(results: list, top_n: int = FINAL_TOP_N) -> list:
    """Return the top N candidates."""
    return results[:top_n]


def results_to_submission_csv(results: list, output_path: str, top_n: int = FINAL_TOP_N):
    """Export top N results to submission-format CSV."""
    top = get_top_candidates(results, top_n)

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["candidate_id", "rank", "score", "reasoning"])
        for i, r in enumerate(top):
            writer.writerow([
                r["candidate_id"],
                i + 1,
                f"{r['final_score']:.6f}",
                r["reasoning"],
            ])

    print(f"\n✅ Submission CSV saved: {output_path} ({len(top)} candidates)")


def results_to_xlsx(results: list, output_path: str, top_n: int = FINAL_TOP_N):
    """Export top N results to XLSX with detailed breakdown."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    top = get_top_candidates(results, top_n)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top Candidates"

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    headers = [
        "Rank", "Candidate ID", "Name", "Current Title", "Years Exp",
        "Location", "Country", "Company", "Industry",
        "Final Score", "Semantic Sim", "Rerank Score",
        "Disqualifier", "Honeypot",
        "Required Skills", "NLP/IR Skills",
        "Response Rate", "GitHub", "Notice (days)",
        "Reasoning"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_idx, r in enumerate(top, 2):
        f = r.get("features", {})
        row_data = [
            r.get("rank", row_idx - 1),
            r.get("candidate_id", ""),
            r.get("name", ""),
            r.get("current_title", ""),
            r.get("years_experience", 0),
            r.get("location", ""),
            r.get("country", ""),
            r.get("company", ""),
            r.get("industry", ""),
            r.get("final_score", 0),
            r.get("semantic_similarity", 0),
            r.get("raw_rerank_score", 0),
            r.get("disqualifier_multiplier", 1.0),
            "⚠️ YES" if r.get("is_honeypot", False) else "No",
            f.get("required_skills_matched", 0),
            f.get("nlp_ir_skill_count", 0),
            f.get("recruiter_response_rate", 0),
            f.get("github_activity", 0),
            f.get("notice_period_days", 0),
            r.get("reasoning", ""),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=(col == len(row_data)))

            # Color-code scores
            if col in (10, 11, 12) and isinstance(val, (int, float)):
                if val >= 0.7:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif val >= 0.4:
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            if col == 14 and "YES" in str(val):
                cell.font = Font(color='FF0000', bold=True)

        if row_idx % 2 == 0:
            for col in range(1, 10):
                cell = ws.cell(row=row_idx, column=col)
                if not cell.fill or cell.fill.start_color.rgb == '00000000':
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    col_widths = [6, 16, 18, 24, 8, 18, 10, 20, 16, 10, 10, 10, 10, 10, 10, 10, 10, 8, 8, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"✅ XLSX report saved: {output_path}")
