"""
Redrob AI Candidate Ranker — Streamlit Application (Neural Edition)
Upload candidates → AI ranks → Download top 100
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json
import time
import os
import sys
import io
import tempfile
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import JOB_DESCRIPTION, FINAL_TOP_N, RERANK_WEIGHTS
from ranking_engine import rank_candidates, get_top_candidates

# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(
    page_title="Redrob AI · Neural Candidate Ranker",
    page_icon="🧠",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ============================================================================
# CUSTOM CSS
# ============================================================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

    .stApp { font-family: 'Inter', sans-serif; }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    .hero-banner {
        background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
        border-radius: 20px;
        padding: 40px 50px;
        margin-bottom: 30px;
        position: relative;
        overflow: hidden;
        border: 1px solid rgba(255,255,255,0.08);
    }
    .hero-banner::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -20%;
        width: 400px;
        height: 400px;
        background: radial-gradient(circle, rgba(99, 102, 241, 0.15) 0%, transparent 70%);
        border-radius: 50%;
    }
    .hero-title {
        font-size: 2.6rem;
        font-weight: 800;
        background: linear-gradient(120deg, #ffffff 0%, #a78bfa 50%, #818cf8 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 8px;
        position: relative;
        z-index: 1;
    }
    .hero-subtitle {
        font-size: 1.1rem;
        color: rgba(255,255,255,0.55);
        font-weight: 400;
        line-height: 1.6;
        position: relative;
        z-index: 1;
    }
    .hero-badge {
        display: inline-block;
        background: linear-gradient(135deg, #6366f1, #8b5cf6);
        color: white;
        padding: 4px 14px;
        border-radius: 50px;
        font-size: 0.75rem;
        font-weight: 600;
        margin-bottom: 12px;
        letter-spacing: 1px;
        position: relative;
        z-index: 1;
    }

    .metric-card {
        background: linear-gradient(145deg, #1e1e2e, #2a2a3e);
        border: 1px solid rgba(255,255,255,0.06);
        border-radius: 16px;
        padding: 24px;
        text-align: center;
        transition: all 0.3s ease;
    }
    .metric-card:hover {
        transform: translateY(-2px);
        border-color: rgba(99, 102, 241, 0.3);
        box-shadow: 0 8px 30px rgba(99, 102, 241, 0.1);
    }
    .metric-value {
        font-size: 2.2rem;
        font-weight: 800;
        background: linear-gradient(120deg, #818cf8, #a78bfa);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .metric-label {
        font-size: 0.82rem;
        color: rgba(255,255,255,0.5);
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 1.5px;
        margin-top: 4px;
    }

    .section-header {
        font-size: 1.4rem;
        font-weight: 700;
        color: #e2e8f0;
        margin: 28px 0 14px 0;
        padding-bottom: 8px;
        border-bottom: 2px solid rgba(99, 102, 241, 0.3);
    }

    .candidate-card {
        background: linear-gradient(145deg, #1a1a2e, #252540);
        border: 1px solid rgba(255,255,255,0.06);
        border-radius: 16px;
        padding: 22px 28px;
        margin-bottom: 12px;
        transition: all 0.3s ease;
    }
    .candidate-card:hover {
        border-color: rgba(99, 102, 241, 0.4);
        box-shadow: 0 4px 20px rgba(99, 102, 241, 0.08);
    }
    .candidate-rank {
        font-size: 1.5rem;
        font-weight: 800;
        background: linear-gradient(120deg, #6366f1, #a78bfa);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .candidate-name {
        font-size: 1.1rem;
        font-weight: 600;
        color: #e2e8f0;
    }
    .candidate-title {
        font-size: 0.88rem;
        color: rgba(255,255,255,0.45);
    }
    .score-badge {
        display: inline-block;
        padding: 5px 14px;
        border-radius: 50px;
        font-weight: 700;
        font-size: 0.9rem;
    }
    .score-high { background: rgba(34, 197, 94, 0.15); color: #22c55e; border: 1px solid rgba(34, 197, 94, 0.3); }
    .score-medium { background: rgba(234, 179, 8, 0.15); color: #eab308; border: 1px solid rgba(234, 179, 8, 0.3); }
    .score-low { background: rgba(239, 68, 68, 0.15); color: #ef4444; border: 1px solid rgba(239, 68, 68, 0.3); }

    .gradient-divider {
        height: 2px;
        background: linear-gradient(90deg, transparent, rgba(99, 102, 241, 0.4), transparent);
        margin: 20px 0;
        border: none;
    }

    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #6366f1, #a78bfa) !important;
    }

    .stDownloadButton > button {
        background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 12px 28px !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        transition: all 0.3s ease !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(99, 102, 241, 0.35) !important;
    }

    .stButton > button {
        background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
        color: white !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 10px 24px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
    }
    .stButton > button:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 20px rgba(99, 102, 241, 0.3) !important;
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f0c29 0%, #1a1a2e 100%);
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# SESSION STATE
# ============================================================================
if "results" not in st.session_state:
    st.session_state.results = None
if "total_candidates" not in st.session_state:
    st.session_state.total_candidates = 0
if "ranking_time" not in st.session_state:
    st.session_state.ranking_time = 0

# ============================================================================
# HERO BANNER
# ============================================================================
st.markdown("""
<div class="hero-banner">
    <div class="hero-badge">NEURAL AI RANKING · SENTENCE-TRANSFORMER</div>
    <div class="hero-title">Intelligent Candidate Discovery</div>
    <div class="hero-subtitle">
        Upload candidate data · Semantic embeddings + structured features · AI-ranked top 100 with reasoning
    </div>
</div>
""", unsafe_allow_html=True)

# ============================================================================
# JD SUMMARY (always shown — it's hardcoded from the challenge)
# ============================================================================
with st.expander("📄 Target Role: Senior AI Engineer — Founding Team, Redrob AI", expanded=False):
    col_jd1, col_jd2 = st.columns(2)
    with col_jd1:
        st.markdown(f"**Title:** {JOB_DESCRIPTION['title']}")
        st.markdown(f"**Company:** {JOB_DESCRIPTION['company']} ({JOB_DESCRIPTION['company_stage']})")
        st.markdown(f"**Location:** {JOB_DESCRIPTION['location']}")
        st.markdown(f"**Experience:** {JOB_DESCRIPTION['experience_range']['min']}-{JOB_DESCRIPTION['experience_range']['max']} years (ideal {JOB_DESCRIPTION['ideal_experience']['min']}-{JOB_DESCRIPTION['ideal_experience']['max']})")
    with col_jd2:
        st.markdown("**Must-have:** Embeddings retrieval, Vector DBs, Evaluation frameworks (NDCG/MRR), Strong Python, NLP/IR")
        st.markdown("**Nice-to-have:** LLM fine-tuning (LoRA), Learning-to-rank, HR-tech exposure")
        st.markdown("**Disqualifiers:** Consulting-only, CV/Speech-only, Title-chasers, Pure researchers")

# ============================================================================
# FILE UPLOAD
# ============================================================================
st.markdown('<div class="section-header">📁 Upload Candidate Data</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Upload candidates file (JSONL, JSON, or gzipped JSONL)",
    type=["jsonl", "json", "gz"],
    help="Upload your candidates.jsonl, candidates.jsonl.gz, or sample_candidates.json file",
)

# ============================================================================
# RANKING
# ============================================================================
if uploaded_file is not None:
    st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

    if st.button("🚀 Run Neural Ranking", use_container_width=True, type="primary"):
        with st.spinner("Loading candidates..."):
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{uploaded_file.name.split(".")[-1]}') as tmp:
                    tmp.write(uploaded_file.read())
                    tmp_path = tmp.name

                candidates = []
                if uploaded_file.name.endswith('.gz'):
                    import gzip
                    with gzip.open(tmp_path, "rt", encoding="utf-8") as f:
                        for line in f:
                            line = line.strip()
                            if line:
                                candidates.append(json.loads(line))
                elif uploaded_file.name.endswith('.jsonl'):
                    with open(tmp_path, "r", encoding="utf-8") as f:
                        for line in f:
                            line = line.strip()
                            if line:
                                candidates.append(json.loads(line))
                else:
                    with open(tmp_path, "r", encoding="utf-8") as f:
                        candidates = json.load(f)

                os.unlink(tmp_path)
                st.session_state.total_candidates = len(candidates)

            except Exception as e:
                st.error(f"❌ Error loading file: {e}")
                st.stop()

        st.info(f"📊 Loaded **{st.session_state.total_candidates:,}** candidates. Starting neural ranking pipeline...")

        # Check for pre-computed embeddings
        emb_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "embeddings")
        if os.path.exists(os.path.join(emb_dir, "candidate_embeddings.npy")):
            st.success("✅ Using pre-computed embeddings (fast mode)")
        else:
            emb_dir = None
            st.warning("⚠️ No pre-computed embeddings found. Computing on-the-fly (this may take a while for large datasets).")

        progress_bar = st.progress(0, text="Initializing pipeline...")
        status_text = st.empty()

        start_time = time.time()

        def update_progress(current, total, stage, elapsed):
            pct = current / max(total, 1)
            progress_bar.progress(min(pct, 0.99), text=f"{stage} · {current:,}/{total:,} · {elapsed:.0f}s")

        results = rank_candidates(
            candidates,
            embeddings_dir=emb_dir,
            progress_callback=update_progress,
        )

        elapsed = time.time() - start_time
        st.session_state.ranking_time = elapsed
        st.session_state.results = results

        progress_bar.progress(1.0, text=f"✅ Complete! {len(results):,} candidates ranked in {elapsed:.1f}s")

# ============================================================================
# RESULTS
# ============================================================================
if st.session_state.results is not None:
    results = st.session_state.results
    top_results = get_top_candidates(results, FINAL_TOP_N)

    st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

    # ====================================================================
    # METRICS
    # ====================================================================
    st.markdown('<div class="section-header">📊 Ranking Overview</div>', unsafe_allow_html=True)

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{st.session_state.total_candidates:,}</div>
            <div class="metric-label">Candidates Analyzed</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{len(top_results)}</div>
            <div class="metric-label">Shortlisted</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        avg_score = sum(r["final_score"] for r in top_results) / len(top_results) if top_results else 0
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{avg_score:.3f}</div>
            <div class="metric-label">Avg Score (Top 100)</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        honeypots = sum(1 for r in top_results if r.get("is_honeypot", False))
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{honeypots}</div>
            <div class="metric-label">Honeypots in Top 100</div>
        </div>
        """, unsafe_allow_html=True)

    with col5:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-value">{st.session_state.ranking_time:.1f}s</div>
            <div class="metric-label">Total Time</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

    # ====================================================================
    # DOWNLOADS
    # ====================================================================
    st.markdown('<div class="section-header">📥 Download Results</div>', unsafe_allow_html=True)

    col_dl1, col_dl2, col_dl3 = st.columns(3)

    with col_dl1:
        # XLSX
        xlsx_buffer = io.BytesIO()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Top 100 Candidates"

            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC'),
            )

            headers = [
                "Rank", "Candidate ID", "Name", "Title", "Years Exp",
                "Location", "Country", "Company",
                "Final Score", "Semantic Sim", "Rerank Score",
                "Honeypot", "Reasoning"
            ]

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            for row_idx, r in enumerate(top_results, 2):
                row_data = [
                    r.get("rank", row_idx - 1),
                    r.get("candidate_id", ""),
                    r.get("name", ""),
                    r.get("current_title", ""),
                    r.get("years_experience", 0),
                    r.get("location", ""),
                    r.get("country", ""),
                    r.get("company", ""),
                    r.get("final_score", 0),
                    r.get("semantic_similarity", 0),
                    r.get("raw_rerank_score", 0),
                    "YES" if r.get("is_honeypot", False) else "No",
                    r.get("reasoning", ""),
                ]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=(col == len(row_data)))
                    if col in (9, 10, 11) and isinstance(val, (int, float)):
                        if val >= 0.7:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        elif val >= 0.4:
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            col_widths = [6, 16, 18, 24, 8, 18, 10, 20, 10, 10, 10, 10, 55]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = 'A2'

            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)

            st.download_button(
                label="📥 Download XLSX Report",
                data=xlsx_buffer,
                file_name=f"ranked_candidates_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"XLSX error: {e}")

    with col_dl2:
        csv_data = "candidate_id,rank,score,reasoning\n"
        for i, r in enumerate(top_results):
            reasoning = r.get("reasoning", "").replace('"', "'").replace(',', ';')
            csv_data += f'{r["candidate_id"]},{i+1},{r["final_score"]:.6f},"{reasoning}"\n'
        st.download_button(
            label="📥 Download CSV Submission",
            data=csv_data,
            file_name=f"submission_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col_dl3:
        # Slim JSON (without features dict for cleaner export)
        slim_results = [{k: v for k, v in r.items() if k != "features"} for r in top_results]
        json_data = json.dumps(slim_results, indent=2, default=str)
        st.download_button(
            label="📥 Download JSON Details",
            data=json_data,
            file_name=f"ranking_details_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
            mime="application/json",
            use_container_width=True,
        )

    st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)

    # ====================================================================
    # TABS
    # ====================================================================
    tab1, tab2, tab3 = st.tabs(["🏆 Ranked Candidates", "📈 Analytics", "🔍 Deep Dive"])

    # ---- TAB 1: TABLE ----
    with tab1:
        st.markdown('<div class="section-header">🏆 Top 100 Ranked Candidates</div>', unsafe_allow_html=True)

        df_data = []
        for r in top_results:
            f = r.get("features", {})
            df_data.append({
                "Rank": r.get("rank", 0),
                "ID": r.get("candidate_id", ""),
                "Name": r.get("name", ""),
                "Title": r.get("current_title", ""),
                "Exp": r.get("years_experience", 0),
                "Location": r.get("location", ""),
                "Company": r.get("company", ""),
                "Score": round(r.get("final_score", 0), 4),
                "Semantic": round(r.get("semantic_similarity", 0), 3),
                "NLP/IR": f.get("nlp_ir_skill_count", 0),
                "AI Skills": f.get("ai_ml_skill_count", 0),
                "Req Match": f.get("required_skills_matched", 0),
                "Response%": f"{f.get('recruiter_response_rate', 0):.0%}",
                "GitHub": f.get("github_activity", 0),
                "Status": "⚠️HP" if r.get("is_honeypot") else ("❌DQ" if r.get("is_disqualified") else "✅"),
                "Reasoning": r.get("reasoning", ""),
            })

        df = pd.DataFrame(df_data)

        def color_scores(val):
            if isinstance(val, float):
                if val >= 0.7:
                    return 'background-color: rgba(34, 197, 94, 0.2); color: #22c55e'
                elif val >= 0.4:
                    return 'background-color: rgba(234, 179, 8, 0.15); color: #eab308'
                else:
                    return 'background-color: rgba(239, 68, 68, 0.15); color: #ef4444'
            return ''

        score_cols = ["Score", "Semantic"]
        if hasattr(df.style, 'map'):
            styled_df = df.style.map(color_scores, subset=score_cols).format(
                {col: "{:.3f}" for col in score_cols}
            )
        else:
            styled_df = df.style.applymap(color_scores, subset=score_cols).format(
                {col: "{:.3f}" for col in score_cols}
            )

        st.dataframe(styled_df, use_container_width=True, height=600, hide_index=True)

    # ---- TAB 2: ANALYTICS ----
    with tab2:
        st.markdown('<div class="section-header">📈 Analytics & Insights</div>', unsafe_allow_html=True)

        col_c1, col_c2 = st.columns(2)

        with col_c1:
            fig = px.histogram(
                df, x="Score", nbins=20,
                title="Score Distribution",
                color_discrete_sequence=["#818cf8"],
                template="plotly_dark",
            )
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Inter", color="#e2e8f0"),
                title_font=dict(size=16, color="#a78bfa"),
            )
            st.plotly_chart(fig, use_container_width=True)

        with col_c2:
            fig2 = px.histogram(
                df, x="Exp", nbins=15,
                title="Experience Distribution",
                color_discrete_sequence=["#a78bfa"],
                template="plotly_dark",
            )
            fig2.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Inter", color="#e2e8f0"),
                title_font=dict(size=16, color="#a78bfa"),
            )
            st.plotly_chart(fig2, use_container_width=True)

        col_c3, col_c4 = st.columns(2)

        with col_c3:
            title_counts = df["Title"].value_counts().head(10)
            fig3 = px.pie(
                names=title_counts.index, values=title_counts.values,
                title="Top Titles",
                template="plotly_dark", color_discrete_sequence=px.colors.sequential.Purp,
            )
            fig3.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Inter", color="#e2e8f0"),
                title_font=dict(size=16, color="#a78bfa"),
            )
            st.plotly_chart(fig3, use_container_width=True)

        with col_c4:
            loc_counts = df["Location"].value_counts().head(10)
            fig4 = px.bar(
                x=loc_counts.values, y=loc_counts.index, orientation='h',
                title="Top Locations",
                template="plotly_dark", color_discrete_sequence=["#818cf8"],
            )
            fig4.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Inter", color="#e2e8f0"),
                title_font=dict(size=16, color="#a78bfa"),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(fig4, use_container_width=True)

        # Radar for top 5
        st.markdown("#### 🎯 Top 5 — Semantic vs Structured Comparison")
        fig_radar = go.Figure()
        categories = ["Semantic Sim", "AI Skills", "NLP/IR", "Req Coverage", "Experience Fit"]
        colors = ['#818cf8', '#a78bfa', '#c084fc', '#e879f9', '#f472b6']

        for i, r in enumerate(top_results[:5]):
            f = r.get("features", {})
            values = [
                r.get("semantic_similarity", 0),
                min(1.0, f.get("ai_ml_skill_count", 0) / 10),
                min(1.0, f.get("nlp_ir_skill_count", 0) / 5),
                f.get("required_skill_coverage", 0),
                f.get("exp_in_ideal_range", 0) * 0.7 + f.get("exp_in_acceptable_range", 0) * 0.3,
            ]
            values.append(values[0])
            fig_radar.add_trace(go.Scatterpolar(
                r=values, theta=categories + [categories[0]],
                fill='toself', name=f"#{i+1} {r.get('name', '')}",
                line=dict(color=colors[i]),
            ))
        fig_radar.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 1], gridcolor='rgba(255,255,255,0.1)'), bgcolor='rgba(0,0,0,0)'),
            template="plotly_dark", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
            font=dict(family="Inter", color="#e2e8f0"), showlegend=True,
        )
        st.plotly_chart(fig_radar, use_container_width=True)

    # ---- TAB 3: DEEP DIVE ----
    with tab3:
        st.markdown('<div class="section-header">🔍 Candidate Deep Dive</div>', unsafe_allow_html=True)

        options = {
            f"#{r['rank']} · {r['name']} · {r['current_title']} ({r['final_score']:.4f})": i
            for i, r in enumerate(top_results)
        }
        selected = st.selectbox("Select candidate:", list(options.keys()))

        if selected:
            idx = options[selected]
            r = top_results[idx]
            f = r.get("features", {})

            score_class = "score-high" if r["final_score"] >= 0.7 else ("score-medium" if r["final_score"] >= 0.4 else "score-low")
            st.markdown(f"""
            <div class="candidate-card">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div>
                        <span class="candidate-rank">#{r['rank']}</span>
                        <span class="candidate-name" style="margin-left: 12px;">{r['name']}</span>
                    </div>
                    <span class="score-badge {score_class}">Score: {r['final_score']:.4f}</span>
                </div>
                <div class="candidate-title" style="margin-top: 6px;">
                    {r['current_title']} at {r.get('company', '?')} · {r['years_experience']:.1f} yrs · {r.get('location', '?')}, {r.get('country', '?')}
                </div>
            </div>
            """, unsafe_allow_html=True)

            # Score breakdown
            col_s1, col_s2, col_s3, col_s4 = st.columns(4)
            metrics = [
                ("Semantic Similarity", r.get("semantic_similarity", 0)),
                ("Rerank Score", r.get("raw_rerank_score", 0)),
                ("Disqualifier", r.get("disqualifier_multiplier", 1.0)),
                ("Honeypot Penalty", r.get("honeypot_penalty", 1.0)),
            ]
            for col, (name, val) in zip([col_s1, col_s2, col_s3, col_s4], metrics):
                with col:
                    color = "#22c55e" if val >= 0.7 else ("#eab308" if val >= 0.4 else "#ef4444")
                    st.markdown(f"""
                    <div class="metric-card">
                        <div style="font-size: 1.5rem; font-weight: 700; color: {color};">{val:.3f}</div>
                        <div class="metric-label">{name}</div>
                    </div>
                    """, unsafe_allow_html=True)

            st.markdown("")

            with st.expander("📊 Key Features"):
                feat_display = {
                    "AI/ML Skills": f.get("ai_ml_skill_count", 0),
                    "NLP/IR Skills": f.get("nlp_ir_skill_count", 0),
                    "Required Skills Matched": f.get("required_skills_matched", 0),
                    "Preferred Skills Matched": f.get("preferred_skills_matched", 0),
                    "Avg AI Proficiency": f"{f.get('avg_ai_proficiency', 0):.2f}",
                    "Career AI Title Ratio": f"{f.get('career_ai_title_ratio', 0):.2f}",
                    "Product Experience": "Yes" if f.get("has_product_experience", 0) else "No",
                    "Consulting Only": "⚠️ Yes" if f.get("is_consulting_only", 0) else "No",
                    "Title Chaser": "⚠️ Yes" if f.get("is_title_chaser", 0) else "No",
                    "Recruiter Response Rate": f"{f.get('recruiter_response_rate', 0):.0%}",
                    "GitHub Activity": f.get("github_activity", 0),
                    "Notice Period": f"{f.get('notice_period_days', 0)} days",
                    "Location Score": f"{f.get('location_score', 0):.2f}",
                    "Days Since Active": f.get("days_since_active", 0),
                }
                st.json(feat_display)

            with st.expander("💬 Reasoning"):
                st.markdown(r.get("reasoning", ""))

            if r.get("disqualifier_reasons"):
                with st.expander("⚠️ Disqualifier Flags"):
                    for reason in r["disqualifier_reasons"]:
                        st.warning(f"🚩 {reason}")

            if r.get("honeypot_flags"):
                with st.expander("🕳️ Honeypot Flags"):
                    for flag in r["honeypot_flags"]:
                        st.error(f"🚩 {flag}")

# ============================================================================
# FOOTER
# ============================================================================
st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)
st.markdown("""
<div style="text-align: center; padding: 20px; color: rgba(255,255,255,0.25); font-size: 0.78rem;">
    🧠 Redrob AI Neural Candidate Ranker · Sentence-Transformer Embeddings + Structured Feature Re-ranking<br>
    CPU-Only · No LLM APIs · Privacy First
</div>
""", unsafe_allow_html=True)
